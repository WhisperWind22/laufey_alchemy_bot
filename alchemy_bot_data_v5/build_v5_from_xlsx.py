# -*- coding: utf-8 -*-
"""
build_v5_from_xlsx.py

Генератор датапака (ingredients_v5.json, effect_categories_v5.csv, ingredient_effect_categories_v5.csv)
из Excel-таблиц ингредиентов.

ВНИМАНИЕ:
- Логика чтения заточена под таблицы из проекта "За Гранью" (формат блоков по 3 строки доп. эффектов).
- Категоризация эффектов автогенерируется по ключевым словам и предназначена как "стартовая".
  Для тонкой настройки редактируйте effect_categories_*.csv вручную.

Usage:
    python build_v5_from_xlsx.py --agnes "Copy of Лаборатория Агнес.xlsx" \
                                 --young "10.2 Алхимия_ Молодой Аристократ.xlsx" \
                                 --outdir .

"""

from __future__ import annotations
import argparse
import json
import os
import re
import unicodedata
from dataclasses import dataclass
from typing import List, Optional, Dict, Tuple
import openpyxl
import pandas as pd


@dataclass
class Ingredient:
    code: str
    name: str
    material: str
    main: str
    add1: str
    add2: str
    add3: str
    source: str


def norm_text(s: str) -> str:
    s = str(s).strip()
    s = s.replace("Ё","Е").replace("ё","е")
    s = s.replace("«", '"').replace("»", '"').replace("“", '"').replace("”", '"')
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", " ", s)
    return s


def norm_key(s: str) -> str:
    s = norm_text(s).lower()
    s = s.replace("галюцина", "галлюцина")
    s = s.replace("приводик", "приводит")
    s = s.replace("сноведени", "сновидени")
    s = s.replace("эфори", "эйфори")
    return s


def detect_tier(t: str) -> Optional[str]:
    tl = norm_key(t)
    if "смертель" in tl:
        return "deadly"
    if "сильн" in tl:
        return "strong"
    if "средн" in tl:
        return "medium"
    if "слаб" in tl:
        return "weak"
    return None


def auto_classify(effect: str) -> Tuple[str, str, str]:
    """
    Returns: kind, tier, tags (semicolon-separated)
    """
    t = norm_text(effect)
    tl = norm_key(effect)
    kind = "raw"
    tier = ""
    tags = []

    if "ядовито для женщин" in tl:
        return "gender_toxin", "", "women_sleep;men_intox"

    if "противояд" in tl:
        kind = "antidote"
        tier = detect_tier(t) or ""
        if "от смертельных ядов" in tl:
            tier = "deadly"
        return kind, tier, ""

    if re.search(r"выводит .*яды", tl) or ("убирает" in tl and "яд" in tl):
        return "detox_poison", detect_tier(t) or "", ""

    if "яд" in tl:
        if "кровотеч" in tl:
            return "poison_bleeding", ("medium" if "средн" in tl else (detect_tier(t) or "medium")), "bleeding"
        if "понижает энергию" in tl or "понижение энергии" in tl:
            return "poison_energy_down", ("weak" if "слаб" in tl else (detect_tier(t) or "weak")), "energy_down"
        return "poison", detect_tier(t) or "", ""

    if "кровоостанавлива" in tl or ("останавливает" in tl and "кровотеч" in tl):
        return "stop_bleeding", "", ""

    if "восстанавливает энергию" in tl:
        return "restore_energy", "", ""

    if "бодрост" in tl or "тониз" in tl:
        return "wake", "", ""

    if "невозможно уснуть" in tl:
        return "cant_sleep", "", "insomnia"

    if "снотвор" in tl or ("сон" in tl and ("вводит" in tl or "погружает" in tl or "сна" in tl)):
        return "sleep", "", ""

    if "отрезв" in tl or "выводит из состояния опьянения" in tl:
        return "sobriety", "", ""

    if "опьян" in tl or "эйфори" in tl:
        return "intoxication", "", ""

    if "стойкость к" in tl and "соблазн" in tl:
        return "temptation_resistance", "", ""

    if "клептоман" in tl:
        return "kleptomania", "", ""

    if "любопытная варвара" in tl:
        return "curious_varvara", "", ""

    if ("правд" in tl) and ("говор" in tl or "заставляет" in tl):
        return "truth", "", ""

    if "врать" in tl or "врет" in tl:
        return "lie", "", ""

    if "галлюцина" in tl or "кошмар" in tl:
        return "hallucinations", "", ""

    if ("менталь" in tl and "защит" in tl) or "против ментального воздействия" in tl or "защищает от ментального воздействия" in tl:
        return "mental_protection", "", ""

    if "разрушает наложенное ментальное воздействие" in tl or "снятие менталь" in tl:
        return "mental_cleanse", "", ""

    if "затмевает" in tl and "разум" in tl:
        return "reason_clouding", "", ""

    if "восстанавливает функционирование составляющей части личности" in tl and "разум" in tl:
        return "reason_restore", "", ""

    # Healing patterns
    if ("исцеляет физические раны" in tl) or ("заживляет раны" in tl) or ("лечит физические раны" in tl) or ("восстанавливает хиты" in tl):
        return "healing", "", ""

    return kind, tier, ";".join(tags)


def parse_agnes(path: str) -> List[Ingredient]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    maxr = ws.max_row
    out = []
    for r in range(1, maxr + 1):
        code = ws.cell(r, 6).value
        name = ws.cell(r, 2).value
        if isinstance(code, str) and code.strip() and isinstance(name, str) and name.strip():
            material = ws.cell(r, 3).value
            main = ws.cell(r, 4).value
            add1 = ws.cell(r, 5).value
            add2 = ws.cell(r + 1, 5).value if r + 1 <= maxr else ""
            add3 = ws.cell(r + 2, 5).value if r + 2 <= maxr else ""
            out.append(Ingredient(
                code.strip(),
                norm_text(name),
                norm_text(material or ""),
                norm_text(main or ""),
                norm_text(add1 or ""),
                norm_text(add2 or ""),
                norm_text(add3 or ""),
                os.path.basename(path),
            ))
    return out


def parse_young(path: str) -> List[Ingredient]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    maxr = ws.max_row
    out = []
    for r in range(2, maxr + 1):
        code = ws.cell(r, 5).value
        name = ws.cell(r, 1).value
        if isinstance(code, str) and code.strip() and isinstance(name, str) and name.strip():
            material = ws.cell(r, 2).value
            main = ws.cell(r, 3).value
            add1 = ws.cell(r, 4).value
            add2 = ws.cell(r + 1, 4).value if r + 1 <= maxr else ""
            add3 = ws.cell(r + 2, 4).value if r + 2 <= maxr else ""
            out.append(Ingredient(
                code.strip(),
                norm_text(name),
                norm_text(material or ""),
                norm_text(main or ""),
                norm_text(add1 or ""),
                norm_text(add2 or ""),
                norm_text(add3 or ""),
                os.path.basename(path),
            ))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--agnes", required=True)
    ap.add_argument("--young", required=True)
    ap.add_argument("--outdir", default=".")
    args = ap.parse_args()

    ings = {}
    for ing in parse_agnes(args.agnes):
        ings[ing.code] = ing
    for ing in parse_young(args.young):
        ings.setdefault(ing.code, ing)

    # ingredients_v5.json
    payload = {
        "version": "v5",
        "generated_at": "local",
        "sources": sorted({ing.source for ing in ings.values()}),
        "ingredients": {
            code: {
                "name": ing.name,
                "material": ing.material,
                "main": ing.main,
                "add1": ing.add1,
                "add2": ing.add2,
                "add3": ing.add3,
                "source": ing.source,
            } for code, ing in sorted(ings.items())
        }
    }
    os.makedirs(args.outdir, exist_ok=True)
    with open(os.path.join(args.outdir, "ingredients_v5.json"), "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    # ingredient_effects_v5.csv
    rows = []
    for code, ing in sorted(ings.items()):
        rows.append({"code": code, "name": ing.name, "material": ing.material, "slot": "main", "slot_index": 0, "effect_text": ing.main, "source": ing.source})
        for i, eff in enumerate([ing.add1, ing.add2, ing.add3], start=1):
            rows.append({"code": code, "name": ing.name, "material": ing.material, "slot": f"add{i}", "slot_index": i, "effect_text": eff, "source": ing.source})
    df = pd.DataFrame(rows)
    df.to_csv(os.path.join(args.outdir, "ingredient_effects_v5.csv"), index=False, encoding="utf-8")

    # effect_categories_v5.csv
    uniq = sorted(set(df["effect_text"].astype(str).map(norm_text)))
    cat_rows = []
    for eff in uniq:
        kind, tier, tags = auto_classify(eff)
        cat_rows.append({"effect_text": norm_text(eff), "norm_key": norm_key(eff), "kind": kind, "tier": tier, "tags": tags})
    dfc = pd.DataFrame(cat_rows)
    dfc.to_csv(os.path.join(args.outdir, "effect_categories_v5.csv"), index=False, encoding="utf-8")

    # ingredient_effect_categories_v5.csv
    cats = {r["effect_text"]: r for r in cat_rows}
    df["effect_text_norm"] = df["effect_text"].map(norm_text)
    df["kind"] = df["effect_text_norm"].map(lambda x: cats.get(x, {}).get("kind", "raw"))
    df["tier"] = df["effect_text_norm"].map(lambda x: cats.get(x, {}).get("tier", ""))
    df["tags"] = df["effect_text_norm"].map(lambda x: cats.get(x, {}).get("tags", ""))
    df.drop(columns=["effect_text_norm"]).to_csv(os.path.join(args.outdir, "ingredient_effect_categories_v5.csv"), index=False, encoding="utf-8")

    print("Done. Files written to:", args.outdir)

if __name__ == "__main__":
    main()
